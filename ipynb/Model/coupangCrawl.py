from bs4 import BeautifulSoup as bs
from pathlib import Path
from openpyxl import Workbook
from fake_useragent import UserAgent
from requests.exceptions import RequestException
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
import time
import os
import re
import requests as rq
import json
import math
import sys
import threading


class ChromeDriver:
    def __init__(self) -> None:
        self.set_options()
        self.set_driver()

    def set_options(self) -> None:
        self.options = Options()
        self.options.add_argument("--headless")
        self.options.add_argument("lang=ko_KR")
        self.options.add_argument(
            "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
        )
        self.options.add_argument("--log-level=3")
        self.options.add_experimental_option("detach", True)
        self.options.add_experimental_option("excludeSwitches", ["enable-logging"])

    def set_driver(self) -> None:
        self.driver = webdriver.Chrome(options=self.options)


class Coupang:
    @staticmethod
    def get_product_code(url: str) -> str:
        prod_code: str = url.split("products/")[-1].split("?")[0]
        return prod_code

    @staticmethod
    def get_soup_object(resp: rq.Response) -> bs:
        return bs(resp.text, "html.parser")

    def __del__(self) -> None:
        if self.ch.driver:
            self.ch.driver.quit()

    def __init__(self) -> None:
        self.base_review_url: str = "https://www.coupang.com/vp/product/reviews"
        self.sd = SaveData()
        self.retries = 10
        self.delay = 0.5
        self.headers = {
            "accept": "*/*",
            "accept-encoding": "gzip, deflate, br, zstd",
            "accept-language": "ko,en;q=0.9,en-US;q=0.8",
            "cookie": "_fbp=fb.1.1709172148924.2042270649; gd1=Y; delivery_toggle=false; srp_delivery_toggle=true; _gcl_au=1.1.1542987164.1726856649; PCID=17272706554699560993959; MARKETID=17272706554699560993959; x-coupang-accept-language=ko-KR; sid=862268b00c90426e990014e6fbf5c7af26562a56; x-coupang-target-market=KR; cf_clearance=r2sQqITwCPo3MoS9IP6Z9PK4HOLYe1ynNWfrauGIufY-1731223910-1.2.1.1-fypmZTk0lIfn8iYHCzX2XZYxGZEjRk5aIM8Eii0PfOgdkWkQ7p.i_AWtifLRou_srPIcGIptzj4pekqm_g.hlWFiPqvRBw5GBlCMFvse9WCy6W0v_ozxtn8yT6Vr8DbALzOVn8umBYJtTVWW_VWJ0sSF5ug8m4hneemnDDLUXeVC9nPEvAQsAEch31CaIy13EZ41WCvs1x.d.G8MTeFLJaGmlJ6fF0pyCtS_JBBdyvloE70.0hQSLZUgN_aKtinLmBEYV9111WtEiy6b_X9rAsmkNfUKmyOZhQrZ0bitPJJTBaL.KOe19sw8rjNSjvLBQWEZjA2sG7ClWr.HeG0VxG3IqSRL.TWBAYgmdBOoJ.DgKkMRXtYr2YEr10JrdaLa5zDHJ5OT6L5Wuc.9bbe2oA; session-id=324aeca3-d8cd-4c45-a6f1-10621d390140; bm_sz=8B4246C44D36EDFC15E70A22536E45B1~YAAQJWHKF+ekeyOTAQAAPjaxUxkwWm66AyB3nnLpzE+aqPPDqugGLbLVDwLT9ONfw/LKyaD3+dznGh6zbEG/dCvrZD1RRuDPnnzoG7/UU37+DKGnwz5/PLTfprF2G0d6ZY0C+EGijaUw1Y0oSODTWswOBXdyBJpub9W/60oWnJocJP0qICJhLVaM95fhCv8+CTHMUMm7I9n/cz5LPTopwEG0gi9WXF5vTH0mKJVpgeB0Rmf4L0ICd0wAO6pl5fnpzzVoQ80k7hgHbUJxpwooF/QNUS6CxRM4s+4lnvWFlDC1w2bsBr66WBlq2+WKXJio1R4kZovl+jl2Dmm9n0AFNKF1KSe7dbBLbWP2i2uLBgiyIS4TFKcwm0NwumkF0Z34PbTTBKjHO1mqtwiAP4P/bu3/vg1UkkSn5qXCQShEN6nMpmEf2YlIQ3ViXNayXFH15pss1AL1AVzufLYgjmkYRNxbNrhB6ZMREQWOfzxjroVvmSkYk+yb8JYNoRxhHZZXvR8spplD3mSsK7aRYyss1qshf3KVx2D9jw==~3753010~4474160;",
            "priority": "u=1, i",
            "sec-ch-ua": '"Chromium";v="131", "Not_A Brand";v="24"',
            "sec-ch-ua-mobile": "?0",
            "sec-ch-ua-platform": '"Windows"',
            "sec-fetch-dest": "empty",
            "sec-fetch-mode": "cors",
            "sec-fetch-site": "same-origin",
            "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
        }
        self.ch = ChromeDriver()

    def get_product_info(self, prod_code: str) -> tuple:
        url = f"https://www.coupang.com/vp/products/{prod_code}"
        self.ch.driver.get(url=url)
        WebDriverWait(self.ch.driver, 20).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "body"))
        )
        page_source: str = self.ch.driver.page_source
        soup = bs(page_source, "html.parser")
        return (
            soup.select_one("h1.prod-buy-header__title").text.strip(),
            int(re.sub("[^0-9]", "", soup.select("span.count")[0].text.strip())),
        )

    def start(self) -> None:
        """크롤링 시작"""
        self.sd.create_directory()
        URL: str = self.input_review_url()
        self.headers["Referer"] = URL
        prod_code: str = self.get_product_code(url=URL)

        # 상품 정보 추출
        try:
            self.title, review_count = self.get_product_info(prod_code=prod_code)
        except Exception as e:
            print({"error": f"상품 기본 정보를 불러오는 도중 오류가 발생했습니다.: {e}"})
            sys.exit()

        review_pages: int = self.calculate_total_pages(review_count)

        # 여러 `payload` 설정 (평점 필터링 예시)
        payloads = [
            {"ratings": "1", "max_reviews": 20},  # 평점 1
            {"ratings": "2", "max_reviews": 20},  # 평점 2
            {"ratings": "", "max_reviews": 50},   # 평점 제한 없음
        ]

        # 각 `payload`를 병렬 처리
        threads = []
        for payload in payloads:
            thread = threading.Thread(target=self.fetch_reviews, args=(prod_code, review_pages, payload))
            threads.append(thread)
            thread.start()

        # 모든 스레드 완료 대기
        for thread in threads:
            thread.join()
        


    def fetch_reviews(self, prod_code: str, review_pages: int, payload: dict) -> None:
        """특정 조건의 리뷰를 가져오기"""
        max_reviews = payload["max_reviews"]
        ratings = payload["ratings"]
        reviews_fetched = 0

        for page in range(1, review_pages + 1):
            if reviews_fetched >= max_reviews:
                break  # 최대 리뷰 개수를 가져왔으면 종료

            # payload 생성
            request_payload = {
                "productId": prod_code,
                "page": page,
                "size": 5,
                "sortBy": "ORDER_SCORE_ASC",
                "ratings": ratings,  # 필터링된 평점
                "q": "",
                "viRoleCode": 2,
                "ratingSummary": True,
            }

            # 데이터 가져오기
            reviews_fetched += self.fetch(request_payload)

    def fetch(self, payload: dict) -> int:
        """리뷰 데이터를 실제로 크롤링하는 메서드"""
        now_page: str = payload["page"]
        print(f"\n[INFO] Start crawling page {now_page} ...\n")
        attempt: int = 0
        reviews_collected = 0  # 수집된 리뷰 수 추적

        while attempt < self.retries:
            try:
                resp = rq.get(
                    url=self.base_review_url,
                    headers=self.headers,
                    params=payload,
                    timeout=10,
                )
                soup = bs(resp.text, "html.parser")

                # 리뷰가 포함된 article 요소 확인
                articles = soup.select("article.sdp-review__article__list")
                if not articles:  # 데이터가 없는 경우
                    print(f"[INFO] No articles found on page {now_page}")
                    return reviews_collected

                # Article Boxes 처리
                for article in articles:
                    dict_data = {
                        "title": self.title,
                        "prod_name": article.select_one(
                            "div.sdp-review__article__list__info__product-info__name"
                        ).text.strip() if article.select_one(
                            "div.sdp-review__article__list__info__product-info__name"
                        ) else "-",
                        "review_date": article.select_one(
                            "div.sdp-review__article__list__info__product-info__reg-date"
                        ).text.strip() if article.select_one(
                            "div.sdp-review__article__list__info__product-info__reg-date"
                        ) else "-",
                        "user_name": article.select_one(
                            "span.sdp-review__article__list__info__user__name"
                        ).text.strip() if article.select_one(
                            "span.sdp-review__article__list__info__user__name"
                        ) else "-",
                        "rating": int(article.select_one(
                            "div.sdp-review__article__list__info__product-info__star-orange"
                        ).attrs.get("data-rating", 0)) if article.select_one(
                            "div.sdp-review__article__list__info__product-info__star-orange"
                        ) else 0,
                        "headline": article.select_one(
                            "div.sdp-review__article__list__headline"
                        ).text.strip() if article.select_one(
                            "div.sdp-review__article__list__headline"
                        ) else "등록된 헤드라인이 없습니다",
                        "review_content": re.sub(
                            "[\\n\\t]", "", article.select_one(
                                "div.sdp-review__article__list__review > div"
                            ).text.strip()
                        ) if article.select_one(
                            "div.sdp-review__article__list__review > div"
                        ) else "등록된 리뷰내용이 없습니다",
                        "answer": article.select_one(
                            "span.sdp-review__article__list__survey__row__answer"
                        ).text.strip() if article.select_one(
                            "span.sdp-review__article__list__survey__row__answer"
                        ) else "맛 평가 없음",
                    }
                    self.sd.save(dict_data)
                    reviews_collected += 1
                    print(dict_data)

                time.sleep(1)
                return reviews_collected  # 수집한 리뷰 개수 반환
            except RequestException as e:
                attempt += 1
                print(f"Attempt {attempt}/{self.retries} failed: {e}")
                if attempt < self.retries:
                    time.sleep(self.delay)
                else:
                    print(f"최대 요청 만료! 다시 실행하세요.")


    @staticmethod
    def clear_console() -> None:
        command: str = "clear"
        if os.name in ("nt", "dos"):
            command = "cls"
        os.system(command=command)

    def input_review_url(self) -> str:
        while True:
            self.clear_console()
            review_url: str = input(
                "원하시는 상품의 URL 주소를 입력해주세요\n\nEx)\nhttps://www.coupang.com/vp/products/7335597976?itemId=18741704367&vendorItemId=85873964906&q=%ED%9E%98%EB%82%B4%EB%B0%94+%EC%B4%88%EC%BD%94+%EC%8A%A4%EB%8B%88%EC%BB%A4%EC%A6%88&itemsCount=36&searchId=0c5c84d537bc41d1885266961d853179&rank=2&isAddedCart=\n\n:"
            )
            if not review_url:
                # 터미널 초기화
                self.clear_console()

                print("URL 주소가 입력되지 않았습니다")
                continue
            return review_url

    def calculate_total_pages(self, review_counts: int) -> int:
        reviews_per_page: int = 5
        return int(math.ceil(review_counts / reviews_per_page))


class SaveData:
    def __init__(self) -> None:
        self.wb: Workbook = Workbook()
        self.ws = self.wb.active
        self.ws.append(["이름", "작성일자", "평점", "리뷰 내용"])
        self.row: int = 2
        self.dir_name: str = "Coupang_reviews"
        self.create_directory()

    def create_directory(self) -> None:
        if not os.path.exists(self.dir_name):
            os.makedirs(self.dir_name)

    def save(self, datas: dict[str, str | int]) -> None:
        file_name: str = os.path.join(self.dir_name + ".xlsx")
        self.ws[f"A{self.row}"] = datas["user_name"]
        self.ws[f"B{self.row}"] = datas["review_date"]
        self.ws[f"C{self.row}"] = datas["rating"]
        self.ws[f"D{self.row}"] = datas["review_content"]
        self.row += 1
        self.wb.save(filename=file_name)

    def __del__(self) -> None:
        self.wb.close()


if __name__ == "__main__":
    coupang = Coupang()
    coupang.start()
